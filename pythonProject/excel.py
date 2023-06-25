import requests
import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
import win32com.client as win32

# Extract data from the workbook exported through benevity
'''
extract = requests.get(r"https://moodys-my.sharepoint.com/:x:/r/personal/appiahe_moodys_com/Documents/Report1687273010287.xlsx?d=wda29f667fa04486f9976846103a7832c&csf=1&web=1&e=U5moVD")
data = extract.json()
'''


extracted_data = pd.read_excel(r"C:\Users\appiahe\OneDrive - moodys.com\Desktop\Report1687273010287.xlsx")

# Clean and process data
df = pd.DataFrame(extracted_data)
# remove the denied & queued entries
denied_word = 'Denied'
queued_word = 'Queued'

filter_list = ['Denied', 'Queued']

# Testing purposes
first_filter = df[df['Volunteer Submission Status'].str.contains(denied_word)]

second_filter = first_filter[first_filter['Volunteer Submission Status'].str.contains(queued_word)]

final_df = df[df['Volunteer Submission Status'].isin(filter_list)]
# Testing purposes

# Remove filtered rows from original
df = df[~df['Volunteer Submission Status'].isin(filter_list)]

# df = df[~df['Volunteer Submission Status'].str.contains(queued_word)]

df.drop_duplicates(subset=['Employee ID'], inplace=True)  # inplaces modifies the data frame rather than
# creating a new one


# print statements to test if successful
print('Original DataFrame:')
print(df) # correct data frame
print('\nFiltered DataFrame:')
print(final_df)


# Perform manipulations using pandas
#

# select all from file and go to teamup2023 report and paste all into unique participation YTD sheet

# TODO : can YTD sheet just be the df?

# Update workbook
path = r"C:\Users\appiahe\OneDrive - moodys.com\Desktop\TeamUp 2023 Volunteering Report_Practice_Enoch.xlsx"
# workBook = load_workbook(
    # r'C:\Users\appiahe\OneDrive - moodys.com\Desktop\TeamUp 2023 Volunteering Report_Practice_Enoch.xlsx')
# param is workbook that we are updating

# original_wb = load_workbook(path, read_only=True)
workbook = Workbook(write_only=True)

writer = pd.ExcelWriter(path, engine="openpyxl")
writer.book = workbook
worksheet = "Unique Participation YTD"
writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
df.to_excel(writer, sheet_name=worksheet, index=False)
writer.save()

# Update the pivot tables
pivot_table1 = workbook["Unique Part. Region"]["A5"].pivotTable
pivot_table1.cacheSource.type = "xlDatabase"
pivot_table1.cacheSource.worksheetSource.sheet = workbook["Unique Participation YTD"].title

pivot_table2 = workbook["Unique Participation Entity"]["A4"].pivotTable
pivot_table2.cacheSource.type = "xlDatabase"
pivot_table2.cacheSource.worksheetSource.sheet = workbook["Unique Participation YTD"].title


workbook.save(path)


# need to take dataframe and turn it into pivot tables

'''
# Load the dataframe
df = pd.read_csv("path/to/dataframe.csv")

# Load the workbook
workbook = load_workbook(filename="path/to/workbook.xlsx")

# Overwrite the sheet with the dataframe
writer = pd.ExcelWriter("path/to/workbook.xlsx", engine="openpyxl")
writer.book = workbook
writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
df.to_excel(writer, sheet_name="Sheet1", index=False)
writer.save()

# Update the pivot tables
pivot_table1 = workbook["Sheet2"]["A1"].pivotTable
pivot_table1.cacheSource.type = "xlDatabase"
pivot_table1.cacheSource.worksheetSource.sheet = workbook["Sheet1"].title

pivot_table2 = workbook["Sheet3"]["A1"].pivotTable
pivot_table2.cacheSource.type = "xlDatabase"
pivot_table2.cacheSource.worksheetSource.sheet = workbook["Sheet1"].title

workbook.save("path/to/workbook.xlsx")



# or take dataframe and use math to pull numbers from it directly for reporting

ytd_sheet = workBook['Unique Participation YTD']  # param is the sheet of which we are updating the data in quotes

# clear existing data in sheet
ytd_sheet.delete_rows(2, ytd_sheet.max_rows)

# write updated data to sheet
for index, row in df.iterrows():
    ytd_sheet.append(row.tolist())
    
    
'''

# TODO uncomment to try this code
'''
# go to unique part. region sheet and select change data source from pivot table menu and paste all the unique participation YTD sheet
region_sheet = workBook['Unique Part. Region']

# TODO use df instead of YTD_sheet
region_table = pd.pivot_table(ytd_sheet, values='Count of Employee ID', index=['CC REGION', 'CC WORK COUNTRY'])

# then go to unique participation entity and select change data source and paste all unique participation YTD
entity_table = pd.pivot_table(ytd_sheet, values='Count of Employee ID', index=['CC ENTITY'])






# param is workbook that we are updating
workBook.RefreshAll()
workBook.Save()
excel.Quit()

'''


# HAVE TO MAKE SURE VISTAPP participation list and benevity sign ups correspond
# originally used =countifs formula to compare the two lists (make sure it is in text form), 0 means didnt participate
# must check for people with two different addresses and people who participated but did not register & those who registered
# but did not participate, if registration is full, remove those who did not participate and recurse last statement
# at benevity, go to track time select all volunteers and uncheck those who did not participate (got a 0 in countifs)
# make sure to also uncheck those that already have time submitted(have track individual time)...
# in benevity, select all and then modify the reward to no reward, then select all again and approve all
# once approved, go to benevity REPORTING << moodys cooperation << reporting << maia << team up 2023 << teamup 2023 logged hours
# then export as .xlsx, in the first row, select filter (check that column headers are formatted properly)
#  in filters, in column volunteer submission status, filter to denied and queued and select all and delete those submissions
#  (delete row), go remove duplicates using employee id as identifier
# select all from sheet and go to teamup excel and paste all into unique participation YTD sheet
# go to unique part. region sheet and select change data source from pivot table menu and paste all the unique participation YTD sheet
# range, then go to unique participation entity and select change data source and paste all unique participation YTD
